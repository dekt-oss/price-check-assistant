from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import dataclass, replace
from decimal import Decimal, InvalidOperation
from pathlib import Path

from purchase_price.schemas import ProductQuery


class QuoteExtractionError(RuntimeError):
    pass


@dataclass(frozen=True)
class QuoteItem:
    source_sheet: str
    source_row: int
    product_name: str = ""
    manufacturer: str = ""
    model_name: str = ""
    specification: str = ""
    quantity: Decimal | None = None
    unit: str = ""
    unit_price: Decimal | None = None
    total_amount: Decimal | None = None
    vat_status: str = ""
    delivery_condition: str = ""
    installation_condition: str = ""
    option_condition: str = ""
    warranty_condition: str = ""
    maintenance_condition: str = ""
    other_conditions: str = ""


@dataclass(frozen=True)
class QuoteExtractionResult:
    items: tuple[QuoteItem, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _PdfDocumentContext:
    manufacturer: str = ""
    model_name: str = ""
    unit: str = ""
    vat_status: str = ""
    installation_condition: str = ""
    option_condition: str = ""
    warranty_condition: str = ""
    other_conditions: str = ""


_FIELD_ALIASES = {
    "product_name": (
        "품명",
        "제품명",
        "상품명",
        "물품명",
        "품목명",
        "내역",
        "commodity",
        "description",
        "commoditydescription",
        "commoditydescriptions",
    ),
    "manufacturer": (
        "제조사",
        "제조업체",
        "메이커",
        "maker",
        "manufacturer",
        "브랜드",
        "brand",
    ),
    "model_name": (
        "모델",
        "모델명",
        "model",
        "modelname",
        "modelno",
        "modelnumber",
    ),
    "specification": (
        "규격",
        "사양",
        "spec",
        "specification",
    ),
    "quantity": (
        "수량",
        "qty",
        "quantity",
    ),
    "unit": (
        "단위",
        "수량단위",
        "포장단위",
        "unit",
        "uom",
    ),
    "unit_price": (
        "단가",
        "견적단가",
        "공급단가",
        "판매단가",
        "unitprice",
        "price",
    ),
    "total_amount": (
        "금액",
        "합계금액",
        "공급가액",
        "총액",
        "amount",
        "total",
        "totalamount",
    ),
    "vat_status": (
        "vat",
        "vat여부",
        "vat포함",
        "vat포함여부",
        "부가세",
        "부가세여부",
        "부가세포함",
        "부가세포함여부",
        "세금",
    ),
    "delivery_condition": (
        "배송",
        "배송비",
        "배송조건",
        "운송",
        "운송비",
        "납품조건",
    ),
    "installation_condition": (
        "설치",
        "설치비",
        "설치조건",
        "설치비용",
    ),
    "option_condition": (
        "옵션",
        "옵션비",
        "옵션조건",
        "부속품",
        "부속",
        "구성",
        "구성품",
    ),
    "warranty_condition": (
        "보증",
        "보증기간",
        "무상보증",
        "무상보증기간",
        "warranty",
    ),
    "maintenance_condition": (
        "유지보수",
        "유지보수조건",
        "유지관리",
        "서비스계약",
        "maintenance",
    ),
    "other_conditions": (
        "기타조건",
        "기타",
        "조건",
        "비고",
        "특이사항",
        "remark",
        "remarks",
        "note",
    ),
}
_SUMMARY_LABELS = frozenset({"합계", "총계", "소계", "부가세", "vat", "공급가액합계"})
_MAX_HEADER_SCAN_ROWS = 30


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s_./()\-]+", "", str(value).strip().casefold())


_ALIAS_LOOKUP = {
    _normalize_header(alias): field
    for field, aliases in _FIELD_ALIASES.items()
    for alias in aliases
}


def _text(value: object) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value).strip())
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _normalize_vat(value: object) -> str:
    text = _text(value)
    folded = text.casefold()
    if not text:
        return ""
    if "면세" in text:
        return "면세"
    if "별도" in text or "미포함" in text or "excluded" in folded or "exclude" in folded:
        return "별도"
    if "포함" in text or "included" in folded or "include" in folded:
        return "포함"
    return text


def parse_quote_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        text = str(value)
    else:
        text = str(value).strip()
        if not text:
            return None
        text = re.sub(r"[,\s₩원\\]", "", text)
        text = text.strip("()")
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if not result.is_finite():
        return None
    return result


def _header_mapping(values: tuple[object, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        normalized = _normalize_header(value)
        field = _ALIAS_LOOKUP.get(normalized)
        if field is None and normalized.startswith("품명"):
            field = "product_name"
        if field is None and normalized.startswith("규격"):
            field = "specification"
        if field is not None and field not in mapping:
            mapping[field] = index
    return mapping


def _header_is_usable(mapping: dict[str, int]) -> bool:
    identifier_count = sum(
        field in mapping
        for field in ("product_name", "manufacturer", "model_name", "specification")
    )
    has_price = "unit_price" in mapping or "total_amount" in mapping
    return identifier_count >= 1 and has_price


def _cell(values: tuple[object, ...], mapping: dict[str, int], field: str) -> object:
    index = mapping.get(field)
    if index is None or index >= len(values):
        return None
    return values[index]


def _is_summary_row(item: QuoteItem) -> bool:
    label = _normalize_header(item.product_name or item.model_name or item.specification)
    return label in _SUMMARY_LABELS


def _has_meaningful_identity(item: QuoteItem) -> bool:
    for value in (item.product_name, item.manufacturer, item.model_name, item.specification):
        compact = re.sub(r"[^0-9A-Za-z가-힣]", "", value)
        if len(compact) >= 2 and not compact.isdigit():
            return True
    return False


def _extract_sheet_rows(
    sheet_name: str,
    rows: Iterable[tuple[object, ...]],
) -> tuple[list[QuoteItem], str | None]:
    row_iterator = iter(enumerate(rows, start=1))
    header_row_number: int | None = None
    mapping: dict[str, int] = {}

    for row_number, values in row_iterator:
        if row_number > _MAX_HEADER_SCAN_ROWS:
            break
        candidate = _header_mapping(tuple(values))
        if _header_is_usable(candidate):
            header_row_number = row_number
            mapping = candidate
            break

    if header_row_number is None:
        return [], f"{sheet_name}: 품목/가격 헤더를 찾지 못해 건너뜀"

    items: list[QuoteItem] = []
    for row_number, values in row_iterator:
        row = tuple(values)
        item = QuoteItem(
            source_sheet=sheet_name,
            source_row=row_number,
            product_name=_text(_cell(row, mapping, "product_name")),
            manufacturer=_text(_cell(row, mapping, "manufacturer")),
            model_name=_text(_cell(row, mapping, "model_name")),
            specification=_text(_cell(row, mapping, "specification")),
            quantity=parse_quote_decimal(_cell(row, mapping, "quantity")),
            unit=_text(_cell(row, mapping, "unit")),
            unit_price=parse_quote_decimal(_cell(row, mapping, "unit_price")),
            total_amount=parse_quote_decimal(_cell(row, mapping, "total_amount")),
            vat_status=_normalize_vat(_cell(row, mapping, "vat_status")),
            delivery_condition=_text(_cell(row, mapping, "delivery_condition")),
            installation_condition=_text(_cell(row, mapping, "installation_condition")),
            option_condition=_text(_cell(row, mapping, "option_condition")),
            warranty_condition=_text(_cell(row, mapping, "warranty_condition")),
            maintenance_condition=_text(_cell(row, mapping, "maintenance_condition")),
            other_conditions=_text(_cell(row, mapping, "other_conditions")),
        )
        if not _has_meaningful_identity(item):
            continue
        if item.unit_price is None and item.total_amount is None:
            continue
        if _is_summary_row(item):
            continue
        items.append(item)

    return items, None


def extract_excel_quote(path: Path) -> QuoteExtractionResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise QuoteExtractionError(
            "Excel(.xlsx) 지원 모듈 openpyxl을 불러올 수 없습니다. 배포 의존성을 확인하세요."
        ) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise QuoteExtractionError(f"Excel 견적서를 읽을 수 없습니다: {exc}") from exc

    items: list[QuoteItem] = []
    warnings: list[str] = []
    try:
        for sheet in workbook.worksheets:
            sheet_items, warning = _extract_sheet_rows(
                sheet.title,
                (tuple(values) for values in sheet.iter_rows(values_only=True)),
            )
            items.extend(sheet_items)
            if warning:
                warnings.append(warning)
    finally:
        workbook.close()

    if not items:
        warnings.append("자동 추출된 견적 품목이 없습니다. 헤더명과 가격 열을 확인하세요.")
    return QuoteExtractionResult(items=tuple(items), warnings=tuple(warnings))


def extract_legacy_excel_quote(path: Path) -> QuoteExtractionResult:
    try:
        import xlrd
    except ImportError as exc:
        raise QuoteExtractionError(
            "구형 Excel(.xls) 지원 모듈 xlrd를 불러올 수 없습니다. "
            "배포 의존성을 확인하거나 .xlsx 파일로 저장해 다시 업로드하세요."
        ) from exc

    try:
        workbook = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as exc:
        raise QuoteExtractionError(f"구형 Excel(.xls) 견적서를 읽을 수 없습니다: {exc}") from exc

    items: list[QuoteItem] = []
    warnings: list[str] = []
    try:
        for sheet in workbook.sheets():
            sheet_items, warning = _extract_sheet_rows(
                sheet.name,
                (tuple(sheet.row_values(index)) for index in range(sheet.nrows)),
            )
            items.extend(sheet_items)
            if warning:
                warnings.append(warning)
    finally:
        workbook.release_resources()

    if not items:
        warnings.append("자동 추출된 견적 품목이 없습니다. 헤더명과 가격 열을 확인하세요.")
    return QuoteExtractionResult(items=tuple(items), warnings=tuple(warnings))


def _pdf_text_rows(text: str) -> tuple[tuple[object, ...], ...]:
    rows: list[tuple[object, ...]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        cells = tuple(cell.strip() for cell in re.split(r"(?:\t+| {2,})", line) if cell.strip())
        if cells:
            rows.append(cells)
    return tuple(rows)


def _clean_pdf_table(table: list[list[object | None]]) -> tuple[tuple[object, ...], ...]:
    cleaned: list[tuple[object, ...]] = []
    for row in table:
        values = tuple(_text(value) for value in row)
        if any(values):
            cleaned.append(values)
    return tuple(cleaned)


def _extract_pdf_line_candidates(text: str, page_number: int) -> list[QuoteItem]:
    amount = r"(?:₩|\\)?\s*\d{1,3}(?:,\d{3})+(?:\.\d+)?"
    row_pattern = re.compile(
        rf"(?P<product>[A-Za-z가-힣][^\n]{{1,100}}?)\s+"
        rf"(?P<spec>[A-Za-z][A-Za-z0-9._/-]{{1,30}})\s+"
        rf"(?P<qty>\d+(?:\.\d+)?)\s+"
        rf"(?P<unit_price>{amount})\s+"
        rf"(?P<total>{amount})(?:\s+(?P<vat>\(?\s*(?:포함|별도|미포함)\s*\)?))?",
        re.IGNORECASE,
    )
    items: list[QuoteItem] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        match = row_pattern.search(line)
        if not match:
            continue
        item = QuoteItem(
            source_sheet=f"PDF {page_number}페이지",
            source_row=line_number,
            product_name=_text(match.group("product")),
            specification=_text(match.group("spec")),
            quantity=parse_quote_decimal(match.group("qty")),
            unit_price=parse_quote_decimal(match.group("unit_price")),
            total_amount=parse_quote_decimal(match.group("total")),
            vat_status=_normalize_vat(match.group("vat") or ""),
        )
        if _has_meaningful_identity(item) and not _is_summary_row(item):
            items.append(item)
    return items


def _extract_pdf_context(texts: Iterable[str]) -> _PdfDocumentContext:
    document = "\n".join(text for text in texts if text)
    if not document:
        return _PdfDocumentContext()

    manufacturer = ""
    model_name = ""
    unit = ""
    vat_status = ""
    installation = ""
    option = ""
    warranty = ""
    other_parts: list[str] = []

    manufacturer_match = re.search(
        r"(?im)^\s*(?:manufacturer|제조사)\s*[:：]\s*([^\r\n]+)", document
    )
    if manufacturer_match:
        manufacturer = _text(manufacturer_match.group(1)).split("  ", 1)[0]

    model_match = re.search(
        r"(?im)\bmodel\s*[-:：]\s*([A-Za-z0-9][A-Za-z0-9._/-]{1,40})", document
    )
    if model_match:
        model_name = _text(model_match.group(1))
    else:
        model_match = re.search(r"(?im)^\s*모델(?:명)?\s*[:：]\s*([^\r\n]+)", document)
        if model_match:
            model_name = _text(model_match.group(1)).split()[0]

    unit_match = re.search(r"(?im)\b1\s+(set|ea|unit|kit|pcs?)\b", document)
    if unit_match:
        unit = unit_match.group(1)

    if re.search(r"(?is)(?:V\.?\s*A\.?\s*T\.?|VAT)[^\n]{0,30}(?:Included|포함)", document):
        vat_status = "포함"
    elif re.search(
        r"(?is)(?:V\.?\s*A\.?\s*T\.?|VAT)[^\n]{0,30}(?:Excluded|별도|미포함)",
        document,
    ):
        vat_status = "별도"

    installation_match = re.search(
        r"(?im)^.*installation and operation.*provided by Contractor.*$", document
    )
    if installation_match:
        installation = "Contractor 설치·운영 제공"
    else:
        installation_match = re.search(r"(?im)^.*설치.*(?:포함|제공).*$", document)
        if installation_match:
            installation = _text(installation_match.group(0))

    warranty_number = re.search(r"(?i)warranty[^\n]{0,100}?\b(\d+)\s*years?\b", document)
    if warranty_number:
        warranty = f"{warranty_number.group(1)}년"
    else:
        word_to_year = {"one": "1", "two": "2", "three": "3", "four": "4", "five": "5"}
        warranty_word = re.search(
            r"(?i)warranty[^\n]{0,120}?\b(one|two|three|four|five)\s+years?\b",
            document,
        )
        if warranty_word:
            warranty = f"{word_to_year[warranty_word.group(1).casefold()]}년"
        else:
            warranty_korean = re.search(r"(?im)(?:보증|warranty)[^\n]{0,50}?(\d+)\s*년", document)
            if warranty_korean:
                warranty = f"{warranty_korean.group(1)}년"

    payment_match = re.search(r"(?im)결제조건\s*[:：]\s*([^\r\n]+)", document)
    if payment_match:
        other_parts.append(f"결제조건: {_text(payment_match.group(1))}")

    option_match = re.search(r"(?im)기타제안\s*[:：]\s*([^\r\n]+)", document)
    if option_match:
        option = _text(option_match.group(1))

    return _PdfDocumentContext(
        manufacturer=manufacturer,
        model_name=model_name,
        unit=unit,
        vat_status=vat_status,
        installation_condition=installation,
        option_condition=option,
        warranty_condition=warranty,
        other_conditions="; ".join(other_parts),
    )


def _merge_conditions(existing: str, inferred: str) -> str:
    if not existing:
        return inferred
    if not inferred or inferred in existing:
        return existing
    return f"{existing}; {inferred}"


def _apply_pdf_context(items: list[QuoteItem], context: _PdfDocumentContext) -> list[QuoteItem]:
    single = len(items) == 1
    enriched: list[QuoteItem] = []
    for item in items:
        enriched.append(
            replace(
                item,
                manufacturer=item.manufacturer or (context.manufacturer if single else ""),
                model_name=item.model_name or (context.model_name if single else ""),
                unit=item.unit or (context.unit if single else ""),
                vat_status=item.vat_status or context.vat_status,
                installation_condition=(
                    item.installation_condition or context.installation_condition
                ),
                option_condition=item.option_condition or context.option_condition,
                warranty_condition=item.warranty_condition or context.warranty_condition,
                other_conditions=_merge_conditions(item.other_conditions, context.other_conditions),
            )
        )
    return enriched


def _dedupe_quote_items(items: Iterable[QuoteItem]) -> list[QuoteItem]:
    deduped: list[QuoteItem] = []
    seen: set[tuple[object, ...]] = set()
    for item in items:
        key = (
            item.product_name.casefold(),
            item.manufacturer.casefold(),
            item.model_name.casefold(),
            item.specification.casefold(),
            item.quantity,
            item.unit.casefold(),
            item.unit_price,
            item.total_amount,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def _extract_with_pdfplumber(path: Path) -> tuple[list[QuoteItem], list[str], list[str], bool]:
    try:
        import pdfplumber
    except ImportError:
        return [], [], [], False

    items: list[QuoteItem] = []
    texts: list[str] = []
    warnings: list[str] = []
    saw_text = False
    try:
        with pdfplumber.open(path) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                try:
                    text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
                except Exception as exc:
                    warnings.append(f"PDF {page_number}페이지: 좌표 텍스트 추출 실패 ({exc})")
                    text = ""
                texts.append(text)
                saw_text = saw_text or bool(text.strip())

                try:
                    tables = page.extract_tables(
                        table_settings={
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "lines",
                            "snap_tolerance": 3,
                            "join_tolerance": 3,
                            "intersection_tolerance": 5,
                        }
                    ) or []
                except Exception as exc:
                    warnings.append(f"PDF {page_number}페이지: 표 경계 추출 실패 ({exc})")
                    tables = []

                for table_number, table in enumerate(tables, start=1):
                    rows = _clean_pdf_table(table)
                    if not rows:
                        continue
                    table_items, _ = _extract_sheet_rows(
                        f"PDF {page_number}페이지 표{table_number}", rows
                    )
                    items.extend(table_items)
    except Exception as exc:
        warnings.append(f"PDF 좌표 기반 파서 사용 실패 ({exc})")
        return [], [], warnings, False

    return items, texts, warnings, saw_text


def _extract_pypdf_text(path: Path) -> tuple[list[str], list[str], bool]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return [], ["PDF fallback 모듈 pypdf를 불러올 수 없습니다."], False

    try:
        reader = PdfReader(str(path))
    except Exception as exc:
        return [], [f"PDF fallback 파서를 사용할 수 없습니다 ({exc})"], False

    texts: list[str] = []
    warnings: list[str] = []
    saw_text = False
    for page_number, page in enumerate(reader.pages, start=1):
        try:
            try:
                text = page.extract_text(extraction_mode="layout") or ""
            except TypeError:
                text = page.extract_text() or ""
        except Exception as exc:
            warnings.append(f"PDF {page_number}페이지: fallback 텍스트 추출 실패 ({exc})")
            text = ""
        texts.append(text)
        saw_text = saw_text or bool(text.strip())
    return texts, warnings, saw_text


def extract_pdf_quote(path: Path) -> QuoteExtractionResult:
    structured_items, plumber_texts, warnings, plumber_saw_text = _extract_with_pdfplumber(path)
    pypdf_texts, fallback_warnings, pypdf_saw_text = _extract_pypdf_text(path)
    warnings.extend(fallback_warnings)

    all_texts: list[str] = []
    for texts in (plumber_texts, pypdf_texts):
        for text in texts:
            if text and text not in all_texts:
                all_texts.append(text)

    saw_text = plumber_saw_text or pypdf_saw_text
    if not saw_text:
        raise QuoteExtractionError(
            "PDF에 추출 가능한 텍스트 레이어가 없습니다. 스캔 이미지형 PDF로 보이며 현재 단계에서는 "
            "OCR을 자동 실행하지 않습니다. 원본 Excel 또는 텍스트 PDF를 사용하거나 OCR 단계가 필요합니다."
        )

    items = [item for item in structured_items if _has_meaningful_identity(item)]
    used_structured_table = bool(items)

    if not items:
        for page_number, text in enumerate(all_texts, start=1):
            items.extend(_extract_pdf_line_candidates(text, page_number))

    if not items:
        for page_number, text in enumerate(pypdf_texts, start=1):
            page_items, _ = _extract_sheet_rows(
                f"PDF {page_number}페이지",
                _pdf_text_rows(text),
            )
            items.extend(item for item in page_items if _has_meaningful_identity(item))

    items = _dedupe_quote_items(items)
    context = _extract_pdf_context(all_texts)
    items = _apply_pdf_context(items, context)

    if items:
        if used_structured_table:
            warnings.append(
                "PDF 표 경계/좌표 기반으로 품목을 추출하고, 후속 페이지의 제조사·모델·VAT·설치·"
                "보증·기타조건을 문서 전체에서 보강했습니다. 원문과 대조해 확인하세요."
            )
        else:
            warnings.append(
                "PDF 표 경계를 안정적으로 복원하지 못해 텍스트 fallback으로 추출했습니다. "
                "VAT·배송·설치·옵션·보증·유지보수 조건을 반드시 원문과 대조하세요."
            )
    else:
        warnings.append(
            "PDF 텍스트는 읽었지만 의미 있는 품목/가격 행을 식별하지 못했습니다. "
            "세액·합계 같은 숫자를 품목으로 오인하지 않도록 자동 추출을 보류했습니다."
        )

    return QuoteExtractionResult(items=tuple(items), warnings=tuple(warnings))


def extract_quote_file(path: Path) -> QuoteExtractionResult:
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        return extract_excel_quote(path)
    if suffix == ".xls":
        return extract_legacy_excel_quote(path)
    if suffix == ".pdf":
        return extract_pdf_quote(path)
    raise QuoteExtractionError("지원하지 않는 파일 형식입니다. .xlsx/.xls/.pdf만 업로드하세요.")


def quote_item_query(item: QuoteItem) -> ProductQuery:
    return ProductQuery(
        product_name=item.product_name,
        manufacturer=item.manufacturer,
        model_name=item.model_name,
        specification=item.specification,
    )
