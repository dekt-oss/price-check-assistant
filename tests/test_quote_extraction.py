from decimal import Decimal
from pathlib import Path

import pytest
import xlwt
from openpyxl import Workbook

import purchase_price.services.quote_extraction as quote_extraction
from purchase_price.services.quote_extraction import (
    QuoteExtractionError,
    extract_quote_file,
    parse_quote_decimal,
    quote_item_query,
)


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "견적"
    sheet.append(["병원 구매 견적서"])
    sheet.append(
        [
            "품명",
            "제조사",
            "모델명",
            "규격",
            "수량",
            "단가",
            "금액",
            "VAT 포함여부",
            "배송비",
            "설치비",
            "옵션",
            "무상보증기간",
            "유지보수",
            "비고",
        ]
    )
    sheet.append(
        [
            "노트북",
            "삼성전자",
            "NT960XJG-K72AG",
            "16GB / 1TB",
            2,
            "2,500,000원",
            "5,000,000",
            "포함",
            "무료",
            "해당없음",
            "기본구성",
            "1년",
            "별도",
            "납기 2주",
        ]
    )
    sheet.append(["합계", "", "", "", "", "", "5,000,000"])

    ignored = workbook.create_sheet("안내")
    ignored.append(["납기 및 보증 조건"])
    workbook.save(path)


def _write_legacy_workbook(path: Path) -> None:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("견적")
    rows = [
        ["병원 구매 견적서"],
        [
            "품명",
            "제조사",
            "모델명",
            "규격",
            "수량",
            "단가",
            "금액",
            "부가세",
            "설치조건",
            "보증기간",
        ],
        [
            "초음파진단기",
            "예시메디칼",
            "US-100",
            "Console",
            1,
            12000000,
            12000000,
            "별도",
            "포함",
            "2년",
        ],
        ["합계", "", "", "", "", "", 12000000],
    ]
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            sheet.write(row_index, column_index, value)
    workbook.save(str(path))


class _FakePdfPage:
    def __init__(self, text: str) -> None:
        self.text = text

    def extract_text(self, **_: object) -> str:
        return self.text


class _FakePdfReader:
    def __init__(self, pages: list[str]) -> None:
        self.pages = [_FakePdfPage(text) for text in pages]


def test_extract_xlsx_quote_finds_header_conditions_and_skips_summary(tmp_path: Path) -> None:
    path = tmp_path / "quote.xlsx"
    _write_workbook(path)

    result = extract_quote_file(path)

    assert len(result.items) == 1
    item = result.items[0]
    assert item.source_sheet == "견적"
    assert item.source_row == 3
    assert item.product_name == "노트북"
    assert item.manufacturer == "삼성전자"
    assert item.model_name == "NT960XJG-K72AG"
    assert item.specification == "16GB / 1TB"
    assert item.quantity == Decimal("2")
    assert item.unit_price == Decimal("2500000")
    assert item.total_amount == Decimal("5000000")
    assert item.vat_status == "포함"
    assert item.delivery_condition == "무료"
    assert item.installation_condition == "해당없음"
    assert item.option_condition == "기본구성"
    assert item.warranty_condition == "1년"
    assert item.maintenance_condition == "별도"
    assert item.other_conditions == "납기 2주"
    assert any("안내" in warning for warning in result.warnings)


def test_extract_xls_quote_finds_header_conditions_and_skips_summary(tmp_path: Path) -> None:
    path = tmp_path / "quote.xls"
    _write_legacy_workbook(path)

    result = extract_quote_file(path)

    assert len(result.items) == 1
    item = result.items[0]
    assert item.source_sheet == "견적"
    assert item.source_row == 3
    assert item.product_name == "초음파진단기"
    assert item.manufacturer == "예시메디칼"
    assert item.model_name == "US-100"
    assert item.specification == "Console"
    assert item.quantity == Decimal("1.0")
    assert item.unit_price == Decimal("12000000.0")
    assert item.total_amount == Decimal("12000000.0")
    assert item.vat_status == "별도"
    assert item.installation_condition == "포함"
    assert item.warranty_condition == "2년"
    assert item.delivery_condition == ""


def test_extract_text_pdf_quote_uses_layout_columns_and_conditions(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    text = "\n".join(
        [
            "병원 구매 견적서",
            "품명  제조사  모델명  규격  수량  단가  금액  VAT  배송비  설치비  보증기간",
            (
                "초음파진단기  예시메디칼  US-100  Console  1  12000000  12000000  "
                "포함  무료  별도  2년"
            ),
            "합계            12000000",
        ]
    )
    monkeypatch.setattr(
        quote_extraction,
        "PdfReader",
        lambda _: _FakePdfReader([text]),
    )
    path = tmp_path / "quote.pdf"

    result = extract_quote_file(path)

    assert len(result.items) == 1
    item = result.items[0]
    assert item.source_sheet == "PDF 1페이지"
    assert item.source_row == 3
    assert item.product_name == "초음파진단기"
    assert item.manufacturer == "예시메디칼"
    assert item.model_name == "US-100"
    assert item.specification == "Console"
    assert item.quantity == Decimal("1")
    assert item.unit_price == Decimal("12000000")
    assert item.total_amount == Decimal("12000000")
    assert item.vat_status == "포함"
    assert item.delivery_condition == "무료"
    assert item.installation_condition == "별도"
    assert item.warranty_condition == "2년"
    assert any("VAT·배송·설치" in warning for warning in result.warnings)


def test_scanned_pdf_without_text_fails_closed(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setattr(
        quote_extraction,
        "PdfReader",
        lambda _: _FakePdfReader(["", "   "]),
    )
    path = tmp_path / "scan.pdf"

    with pytest.raises(QuoteExtractionError, match="텍스트 레이어"):
        extract_quote_file(path)


def test_quote_item_converts_to_existing_product_query(tmp_path: Path) -> None:
    path = tmp_path / "quote.xlsx"
    _write_workbook(path)
    item = extract_quote_file(path).items[0]

    query = quote_item_query(item)

    assert query.product_name == "노트북"
    assert query.manufacturer == "삼성전자"
    assert query.model_name == "NT960XJG-K72AG"
    assert query.specification == "16GB / 1TB"


def test_rows_without_a_price_are_not_auto_extracted(tmp_path: Path) -> None:
    path = tmp_path / "quote.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["품명", "모델명", "단가", "VAT"])
    sheet.append(["노트북", "NT960XJG-K72AG", None, "포함"])
    workbook.save(path)

    result = extract_quote_file(path)

    assert result.items == ()
    assert any("자동 추출된 견적 품목이 없습니다" in warning for warning in result.warnings)


def test_unknown_condition_headers_do_not_get_invented(tmp_path: Path) -> None:
    path = tmp_path / "quote.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["품명", "모델명", "단가", "납기메모"])
    sheet.append(["노트북", "NT960XJG-K72AG", 1000000, "보증 포함 같음"])
    workbook.save(path)

    item = extract_quote_file(path).items[0]

    assert item.warranty_condition == ""
    assert item.other_conditions == ""


def test_parse_quote_decimal_handles_common_krw_display() -> None:
    assert parse_quote_decimal("₩ 1,234,500원") == Decimal("1234500")
    assert parse_quote_decimal(1200) == Decimal("1200")
    assert parse_quote_decimal("문의") is None


def test_unsupported_file_type_fails_closed(tmp_path: Path) -> None:
    path = tmp_path / "quote.txt"
    path.write_text("not a quote", encoding="utf-8")

    with pytest.raises(QuoteExtractionError, match=r"\.xlsx/\.xls/\.pdf"):
        extract_quote_file(path)
